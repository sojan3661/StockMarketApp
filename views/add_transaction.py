import streamlit as st
import datetime
import io
import sys
import os
import openpyxl
import time

# Add the app root directory to Python path to allow imports from Config
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from Config.supabase_client import db

st.title("Add Transaction")


# Verification check for credentials
if not db.is_configured():
    st.warning("⚠️ Supabase credentials not found!")
    st.info("Please set your credentials directly inside the init method of `Config/supabase_client.py`.")
    st.stop()
    
@st.cache_data(ttl=300)
def load_add_tx_db_data():
    return (
        db.fetch_stocks(),
        db.fetch_investment_plan(),
        db.fetch_open_transactions(),
        db.fetch_allocations(),
        db.fetch_stock_allocations(),
        db.fetch_currency_pairs()
    )

with st.spinner("Loading available assets & portfolios..."):
    (
        db_stocks,
        db_investment_plan,
        open_transactions,
        db_allocations,
        db_stock_allocations,
        db_currency_pairs
    ) = load_add_tx_db_data()
    
# Create a list of available portfolios
if not db_investment_plan:
    available_portfolios = []
else:
    plans_list = db_investment_plan if isinstance(db_investment_plan, list) else [db_investment_plan]
    available_portfolios = sorted([p.get("Portfolio", "") for p in plans_list if p.get("Portfolio", "")])

# Create a list of available symbols from StockManagement
if not db_stocks:
    available_symbols = []
    st.info("No assets found in your portfolio yet. Go to 'Stock Management' to start adding assets before recording transactions.")
else:
    available_symbols = sorted([p.get("Symbol", "") for p in db_stocks if p.get("Symbol", "")])

# -----------------------------
# Currency Conversion Helpers
# -----------------------------
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_exchange_rate(pair_currency, date_str):
    # Try YFinance directly first
    try:
        import yfinance as yf
        dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        end_dt = dt + datetime.timedelta(days=5)
        ticker = yf.Ticker(f"{pair_currency}=X")
        hist = ticker.history(start=dt.strftime("%Y-%m-%d"), end=end_dt.strftime("%Y-%m-%d"))
        if not hist.empty:
            return float(hist.iloc[0]['Close'])
    except Exception as e:
        print(f"Error fetching {pair_currency} via yfinance: {e}")

    # Fallback 1: Yahoo Finance query2 API (Bypasses some Streamlit Cloud blocks)
    try:
        import urllib.request
        import json
        import ssl
        
        dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        period1 = int(dt.timestamp())
        period2 = int((dt + datetime.timedelta(days=5)).timestamp())
        
        url = f"https://query2.finance.yahoo.com/v8/finance/chart/{pair_currency}=X?period1={period1}&period2={period2}&interval=1d"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        context = ssl._create_unverified_context()
        with urllib.request.urlopen(req, context=context, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            res = data.get("chart", {}).get("result")
            if res and res[0].get("indicators", {}).get("quote"):
                closes = res[0]["indicators"]["quote"][0].get("close")
                for c in closes:
                    if c is not None:
                        return float(c)
    except Exception as e:
        print(f"Error fetching {pair_currency} via Yahoo API fallback: {e}")

    # Fallback 2: Frankfurter API (Free, no key required, highly reliable for historical fiat)
    try:
        import urllib.request
        import json
        import ssl
        
        if len(pair_currency) == 6:
            base = pair_currency[:3]
            target = pair_currency[3:]
            url = f"https://api.frankfurter.app/{date_str}?from={base}&to={target}"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            context = ssl._create_unverified_context()
            with urllib.request.urlopen(req, context=context, timeout=5) as response:
                data = json.loads(response.read().decode('utf-8'))
                rates = data.get("rates", {})
                if target in rates:
                    return float(rates[target])
    except Exception as e:
        print(f"Error fetching {pair_currency} via Frankfurter API fallback: {e}")

    return None

def compute_converted_values(symbol, price, date_str, db_stocks, db_currency_pairs):
    country = "INDIA"
    for s in db_stocks:
        if s.get("Symbol") == symbol:
            country = s.get("Country", "INDIA")
            break
            
    local_avg_value = price
    usd_avg_value = price
    
    if country.upper() != "INDIA" and country:
        pair_currency = None
        for cp in db_currency_pairs:
            if cp.get("Country", "").upper() == country.upper():
                pair_currency = cp.get("PairCurrency")
                break
        
        if pair_currency:
            rate = fetch_exchange_rate(pair_currency, date_str)
            if rate is not None:
                local_avg_value = price * rate
                
    usdinr_rate = fetch_exchange_rate("USDINR", date_str)
    if usdinr_rate is not None:
        usd_avg_value = local_avg_value / usdinr_rate
        
    return local_avg_value, usd_avg_value

# -----------------------------
# Bulk Upload Processing
# -----------------------------
@st.cache_data
def generate_transaction_template(portfolios, symbols) -> bytes:
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Headers
    ws.append(["Portfolio", "Symbol", "Type", "Qty", "Avg", "Date"])
    
    # Add a clear note about Date format
    from openpyxl.styles import Font
    ws["G1"] = "← NOTE: Please enter Date in DD-MM-YYYY format (e.g., 25-12-2023)"
    ws["G1"].font = Font(italic=True, color="0000FF") # Blue italic font

    # ── Hidden reference sheet for long dropdown lists ────────────────────────
    ref_ws = wb.create_sheet("_Ref")
    ref_ws.sheet_state = "hidden"

    # Write portfolios to _Ref col A
    for i, p in enumerate(portfolios, start=1):
        ref_ws.cell(row=i, column=1, value=p)
    port_range = f"_Ref!$A$1:$A${max(len(portfolios), 1)}" if portfolios else None

    # Write symbols to _Ref col B
    for i, s in enumerate(symbols, start=1):
        ref_ws.cell(row=i, column=2, value=s)
    sym_range = f"_Ref!$B$1:$B${max(len(symbols), 1)}" if symbols else None

    # ── Portfolio dropdown (col A) ────────────────────────────────────────────
    if port_range:
        dv_port = DataValidation(type="list", formula1=port_range,
                                 allow_blank=True, showDropDown=False)
        dv_port.sqref = "A2:A1000"
        ws.add_data_validation(dv_port)

    # ── Symbol dropdown (col B) ───────────────────────────────────────────────
    if sym_range:
        dv_sym = DataValidation(type="list", formula1=sym_range,
                                allow_blank=True, showDropDown=False)
        dv_sym.sqref = "B2:B1000"
        ws.add_data_validation(dv_sym)

    # ── Type dropdown (col C) ─────────────────────────────────────────────────
    dv_type = DataValidation(type="list", formula1='"Buy,Sell"',
                             allow_blank=True, showDropDown=False)
    dv_type.sqref = "C2:C1000"
    ws.add_data_validation(dv_type)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

col_dl, col_ul, _ = st.columns([1, 1, 2])
with col_dl:
    import base64
    b64_data = base64.b64encode(generate_transaction_template(available_portfolios, available_symbols)).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_data}" download="transaction_template.xlsx" style="display: block; width: 100%; padding: 0.5rem 1rem; background-color: #2D333B; border: 1px solid #4B5563; color: #E2E8F0; text-align: center; text-decoration: none; border-radius: 8px; font-weight: 500; box-sizing: border-box; transition: background-color 0.2s;">📥 Download Template</a>'
    st.markdown(href, unsafe_allow_html=True)

with col_ul:
    uploaded_file = st.file_uploader(
        "Upload Transactions",
        type=["xlsx"],
        label_visibility="collapsed"
    )

# -----------------------------
# Bulk Upload Processing
# -----------------------------
if uploaded_file is not None:
    with st.expander("📋 Preview & Import Uploaded Transactions", expanded=True):
        try:
            import pandas as pd
            upload_df = pd.read_excel(uploaded_file, dtype=str)

            # Normalise column names (strip spaces, title-case)
            upload_df.columns = [c.strip() for c in upload_df.columns]

            required_cols = {"Portfolio", "Symbol", "Type", "Qty", "Avg", "Date"}
            missing = required_cols - set(upload_df.columns)
            if missing:
                st.error(f"Missing columns in file: {', '.join(missing)}")
            else:
                # Build a display copy with Date formatted as dd-mmm-yyyy
                display_df = upload_df.copy()
                try:
                    display_df["Date"] = pd.to_datetime(
                        upload_df["Date"], dayfirst=True
                    ).dt.strftime("%d-%b-%Y")
                except Exception:
                    pass  # Leave as-is if parsing fails

                st.dataframe(display_df, use_container_width=True, hide_index=True)

                if st.button("🚀 Import All Transactions", type="primary"):
                    success_count = 0
                    fail_count = 0

                    for i, row in upload_df.iterrows():
                        port  = str(row["Portfolio"]).strip()
                        sym   = str(row["Symbol"]).strip()
                        ttype = str(row["Type"]).strip().capitalize()
                        qty   = float(row["Qty"])
                        avg   = float(row["Avg"])
                        
                        raw_date = str(row["Date"]).strip()
                        try:
                            # Parse date safely (assuming dayfirst for Indian format / common input)
                            # and convert to Supabase expected ISO format YYYY-MM-DD
                            date = pd.to_datetime(raw_date, dayfirst=True).strftime("%Y-%m-%d")
                        except Exception:
                            # Fallback if unparseable, though this might cause DB error
                            date = raw_date[:10]

                        if ttype == "Buy":
                            local_val, usd_val = compute_converted_values(sym, avg, date, db_stocks, db_currency_pairs)
                            ok = db.add_buy_transaction(
                                symbol=sym,
                                quantity=qty,
                                price=avg,
                                date=date,
                                portfolio=port,
                                buy_avg_local=local_val,
                                buy_avg_usd=usd_val
                            )
                        elif ttype == "Sell":
                            local_val, usd_val = compute_converted_values(sym, avg, date, db_stocks, db_currency_pairs)
                            ok = db.process_sell_transaction(
                                symbol=sym,
                                sell_qty=qty,
                                sell_avg=avg,
                                sell_date=date,
                                portfolio=port,
                                sell_avg_local=local_val,
                                sell_avg_usd=usd_val
                            )
                        else:
                            st.warning(f"Row {i+2}: Unknown Type '{ttype}' — skipped.")
                            fail_count += 1
                            continue

                        if ok:
                            success_count += 1
                        else:
                            fail_count += 1

                    if success_count:
                        st.success(f"✅ {success_count} transaction(s) imported successfully.")
                    if fail_count:
                        st.error(f"❌ {fail_count} transaction(s) failed — see errors above.")

        except Exception as e:
            st.error(f"Error reading file: {e}")

st.markdown(
    """
    <div style="background-color: #1E222A; padding: 15px 20px; border-radius: 8px 8px 0 0; border: 1px solid #2D333B; border-bottom: none; margin-top: 20px;">
        <h3 style="margin: 0; color: #F8FAFC; font-size: 1.2rem;">📝 Transaction Details</h3>
    </div>
    """,
    unsafe_allow_html=True
)

# Use a standard container instead of a form to allow dynamic UI updates like Sell All
if "tx_form_key" not in st.session_state:
    st.session_state["tx_form_key"] = 0
    
container = st.container()

@st.dialog("Mutual Fund Calculator")
def mf_calc_dialog(q_key, p_key, tx_type):
    st.write(f"Enter the **Quantity** and the **Total {tx_type} Amount**:")
    
    current_q = float(st.session_state.get(q_key, 1.0))
    current_p = float(st.session_state.get(p_key, 100.0))
    current_total = current_q * current_p
    
    calc_qty = st.number_input("Quantity", min_value=0.0001, value=current_q, format="%.4f", step=1.0)
    calc_total = st.number_input(f"Total {tx_type} Amount (₹)", min_value=0.01, value=current_total, step=500.0, format="%.2f")
    
    if st.button("OK", type="primary"):
        st.session_state[q_key] = float(calc_qty)
        st.session_state[p_key] = float(calc_total / calc_qty) if calc_qty > 0 else 0.0
        st.rerun()



with container:
    col1, col2 = st.columns(2)

    with col1:
        default_port = st.session_state.get("last_selected_portfolio")
        port_index = 0
        if default_port and default_port in available_portfolios:
            port_index = available_portfolios.index(default_port)

        selected_portfolio = st.selectbox(
            "Portfolio", 
            options=available_portfolios,
            index=port_index,
            help="Select the portfolio this transaction belongs to.",
            key=f"port_{st.session_state['tx_form_key']}"
        )
        
        if selected_portfolio:
            allocated_symbols = {
                a.get("Symbol") for a in (db_stock_allocations or [])
                if a.get("Portfolio") == selected_portfolio and float(a.get("Allocation") or 0) > 0 and a.get("Symbol")
            }
            open_tx_symbols = {
                tx.get("Symbol") for tx in (open_transactions or [])
                if tx.get("Portfolio") == selected_portfolio and tx.get("Symbol")
            }
            valid_symbols = allocated_symbols.union(open_tx_symbols)
            filtered_symbols = sorted([
                s for s in available_symbols
                if s in valid_symbols
            ])
            options_to_show = filtered_symbols if filtered_symbols else available_symbols
        else:
            options_to_show = available_symbols

        selected_symbol = st.selectbox(
            "Asset Symbol", 
            options=options_to_show,
            help="Select an asset you've already added via Stock Management.",
            key=f"sym_{st.session_state['tx_form_key']}"
        )

    with col2:
        transaction_type = st.radio(
            "Transaction Type",
            options=["Buy", "Sell"],
            horizontal=True,
            key=f"type_{st.session_state['tx_form_key']}"
        )
        
        sell_all = False
        current_qty = 0.0
        if transaction_type == "Sell" and open_transactions:
            current_qty = sum(
                float(tx.get("Qty", 0))
                for tx in open_transactions
                if tx.get("Portfolio") == selected_portfolio and tx.get("Symbol") == selected_symbol
            )
            if current_qty > 0:
                sell_all = st.checkbox(f"Sell All ({current_qty:.4f} available)", key=f"sell_{st.session_state['tx_form_key']}")

    with col1:
        qty_key = f"qty_{st.session_state['tx_form_key']}"
        price_key = f"price_{st.session_state['tx_form_key']}"
        
        if qty_key not in st.session_state:
            st.session_state[qty_key] = 1.0
        if price_key not in st.session_state:
            st.session_state[price_key] = 100.0
            
        is_mf = any(p.get("Symbol") == selected_symbol and not p.get("Equity", True) for p in db_stocks)
        if is_mf:
            if st.button("🧮 Calculate from Total Amount"):
                mf_calc_dialog(qty_key, price_key, transaction_type)
            
        if sell_all and current_qty > 0:
            st.session_state[qty_key] = float(current_qty)
            
        quantity = st.number_input(
            "Quantity",
            min_value=0.0001, # Allows partial shares/units
            step=1.0,
            format="%.4f",
            help="Number of shares or mutual fund units.",
            disabled=sell_all,
            key=qty_key
        )
        
        default_date = st.session_state.get("last_transaction_date", datetime.date.today())
        transaction_date = st.date_input(
            "Transaction Date",
            value=default_date,
            key=f"date_{st.session_state['tx_form_key']}"
        )
        
    with col2:
        price = st.number_input(
            "Price per unit",
            min_value=0.01,
            step=1.0,
            format="%.2f",
            help="The price at which the asset was bought or sold.",
            key=f"price_{st.session_state['tx_form_key']}"
        )
        
        loc_key = f"local_val_{st.session_state['tx_form_key']}"
        usd_key = f"usd_val_{st.session_state['tx_form_key']}"
        if loc_key in st.session_state:
            st.success(f"Local Value: ₹{st.session_state[loc_key]:.2f} | USD Value: ${st.session_state[usd_key]:.2f}")

    # Submission
    submitted = st.button("💾 Save Transaction", type="primary", use_container_width=True)

if submitted:
    if not available_symbols:
        st.error("Cannot add transaction without an available asset symbol.")
    elif not selected_symbol:
        st.error("Please select an asset symbol.")
    else:
        # Format the date properly for the DB as YYYY-MM-DD
        formatted_date = transaction_date.strftime("%Y-%m-%d")
        
        with st.spinner(f"Processing {transaction_type} transaction..."):
            loc_key = f"local_val_{st.session_state['tx_form_key']}"
            usd_key = f"usd_val_{st.session_state['tx_form_key']}"
            
            if loc_key in st.session_state and usd_key in st.session_state:
                local_val = st.session_state.get(loc_key)
                usd_val = st.session_state.get(usd_key)
            else:
                local_val, usd_val = compute_converted_values(selected_symbol, float(price), formatted_date, db_stocks, db_currency_pairs)

            if transaction_type == "Buy":
                success = db.add_buy_transaction(
                    symbol=selected_symbol,
                    quantity=float(quantity),
                    price=float(price),
                    date=formatted_date,
                    portfolio=selected_portfolio,
                    buy_avg_local=local_val,
                    buy_avg_usd=usd_val
                )
            else:
                success = db.process_sell_transaction(
                    symbol=selected_symbol,
                    sell_qty=float(quantity),
                    sell_avg=float(price),
                    sell_date=formatted_date,
                    portfolio=selected_portfolio,
                    sell_avg_local=local_val,
                    sell_avg_usd=usd_val
                )
            
        if success:
            st.session_state["last_selected_portfolio"] = selected_portfolio
            st.session_state["last_transaction_date"] = transaction_date
            st.success(f"Successfully recorded {transaction_type} of {quantity} {selected_symbol} @ ₹{price}")
            time.sleep(1.5)
            st.session_state["tx_form_key"] += 1
            st.rerun()
